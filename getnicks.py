import openpyxl
import shutil


def returnnicks():
    fname = r'C:\Users\Cat Behemoth\OneDrive\walkdead\cards.xlsx'
    # shutil.copyfile(fname, 'cards.xlsx')
    # shutil.copystat(fname, 'cards.xlsx')
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name('nicks')
    nicks = list()
    i = 1
    while True:
        if sheet.cell(row=i, column=1).value != None:
            nicks.append(sheet.cell(row=i, column=1).value)
            i += 1
        else:
            break
    return nicks

def putinxl(nick, wk, rd, mp, mc, sf, sc, ph, pw, cc, sr, lv, i):
