import psycopg2
import openpyxl
import getpass

#connect to base
passwd = getpass.getpass("Enter password:", )
connection = psycopg2.connect(dbname="TWD", user="postgres", password=passwd, host="194.67.105.168")
cursor = connection.cursor()
#read existent players
cursor.execute('SELECT * from "players"')
players = cursor.fetchall()
cursor.close()
connection.close()
maxid = 0
playersdb= list()
if len(players) != 0:
    for i in range(len(players)):
        if players[i][0] > maxid:
            maxid = players[i][0]
        playersdb.append(players[i][1])

#read players from excel
fname = r'C:\Users\Cat Behemoth\OneDrive\walkdead\cards.xlsx'
wb = openpyxl.load_workbook(fname)
sheet = wb['nicks']
playersxl = list()
i = 1
while sheet.cell(row=i, column=1).value != None:
    playersxl.append(sheet.cell(row=i, column=1).value)
    i += 1
wb.close()

#create list to add
playersput = list()
for i in range(len(playersxl)):
    if playersxl[i] not in playersdb:
        playersput.append(playersxl[i])

#writing to server if needed
if len(playersput) > 0:
    connection = psycopg2.connect(dbname="RFR", user="postgres", password=passwd, host="194.67.105.168")
    cursor = connection.cursor()
    for i in range(len(playersput)):
        idd = maxid + 1 + i
        cursor.execute('INSERT INTO "players" ("Id", "Name") VALUES (' + str(idd) + " ,'" + playersput[i] + "')")
    connection.commit()
    connection.close()