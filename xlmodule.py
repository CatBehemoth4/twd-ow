import openpyxl, openpyxl.utils, openpyxl.styles


#getting the list of possible nicks
def returnnicks():
    fname = r'C:\Users\Cat Behemoth\Documents\wrks\walkdead\cards.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name('nicks')
    nicks = list()
    i = 1
    while True:
        if sheet.cell(row=i, column=1).value != None:
            nicks.append(sheet.cell(row=i, column=1).value)
            i += 1
        else:
            break
    return nicks

def putinxl(data):
#main file location
    fname = r'C:\Users\Cat Behemoth\Documents\wrks\walkdead\cards.xlsx'
    wb = openpyxl.load_workbook(filename=fname)
    sheet = wb.get_sheet_by_name('Run, Forest, Run')
#getting number of previous inpuys
    numbers = sheet.cell(row=1, column=3).value
#iterating each nicks
    for i in range(len(data)):
#starting position (swipe right 15 columns per each previous input)
        col = numbers * 15 + 1
        row = i + 3
#writing data
        sheet.cell(row=row, column=col).value = data[i][0]
        sheet.cell(row=row, column=col + 1).value = data[i][1]
        sheet.cell(row=row, column=col + 2).value = data[i][2]
        sheet.cell(row=row, column=col + 3).value = data[i][3]
        sheet.cell(row=row, column=col + 4).value = data[i][4]
        sheet.cell(row=row, column=col + 5).value = data[i][5]
        sheet.cell(row=row, column=col + 6).value = data[i][6]
        sheet.cell(row=row, column=col + 7).value = data[i][7]
        sheet.cell(row=row, column=col + 8).value = data[i][8]
        sheet.cell(row=row, column=col + 9).value = data[i][9]
        sheet.cell(row=row, column=col + 10).value = data[i][10]
        sheet.cell(row=row, column=col + 11).value = data[i][11]
# formula for shots per kill
        firstcol = openpyxl.utils.get_column_letter(col + 5)
        secndcol = openpyxl.utils.get_column_letter(col + 1)
        formula = '=' + firstcol + str(row) + '/' + secndcol + str(row)
        sheet.cell(row=row, column=col + 12).value = formula
#formula for % of successive missions
        firstcol = openpyxl.utils.get_column_letter(col + 4)
        secndcol = openpyxl.utils.get_column_letter(col + 3)
        formula = '=' + firstcol + str(row) + '/' + secndcol + str(row)
        sheet.cell(row=row, column=col + 13).value = formula
#update start position for next input
    sheet.cell(row=1, column=3).value += 1
#apply styles for nicks
    firstcol = openpyxl.utils.get_column_letter((numbers - 1) * 15 + 1)
    cellname1 = firstcol + '3'
    firstcol = openpyxl.utils.get_column_letter(numbers * 15 + 1)
    for i in range (3, 31):
        cellname = firstcol + str(i)
        sheet[cellname]._style = sheet[cellname1]._style
# apply styles for inside-game stats
    firstcol = openpyxl.utils.get_column_letter((numbers - 1) * 15 + 2)
    cellname1 = firstcol + '3'
    for i in range (2, 13):
        firstcol = openpyxl.utils.get_column_letter(numbers * 15 + i)
        for j in range(27):
            cellname = firstcol + str(j + 3)
            sheet[cellname]._style = sheet[cellname1]._style
# apply styles for additional stats
    firstcol = openpyxl.utils.get_column_letter((numbers - 1) * 15 + 13)
    cellname1 = firstcol + '3'
    for i in range(13, 15):
        firstcol = openpyxl.utils.get_column_letter(numbers * 15 + i)
        for j in range(27):
            cellname = firstcol + str(j + 3)
            sheet[cellname]._style = sheet[cellname1]._style
#save current input
    wb.save(filename=fname)
    wb.close()


def putinxlext(data):
    # main file location
    fname = r'C:\Users\Cat Behemoth\Documents\wrks\walkdead\cards.xlsx'
    wb = openpyxl.load_workbook(filename=fname)
    sheet = wb.get_sheet_by_name('other')
    # initial row
    row = 2
    # iterating each nicks
    for plr in data:
        col = 1
        for val in plr:
            sheet.cell(row=row, column=col).value = val
            col += 1
        row += 1
    # save current input
    wb.save(filename=fname)
    wb.close()